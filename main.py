from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

def append_transaction(file_path, date, description, category, amount, transaction_type, recurring, note=""):
    # Load workbook and active sheet
    try: 
        workbook = load_workbook(file_path)
        # Selecting the active sheet
        worksheet = workbook.active

        # List of transation data to add
        transaction_data = [date, description, category, amount, transaction_type, recurring, note]
        
        # Append the transaction data to the next row
        ## @todo May need to insert into the correct place or see if I can sort it
        worksheet.append(transaction_data)


        last_row = worksheet.max_row

        date_cell = worksheet[f"A{last_row}"]
        date_cell.number_format = "DD/MM/YYYY"

        amount_cell = worksheet[f"D{last_row}"]
        amount_cell.number_format = "£#,##0.00"



        # Save the updated workbook
        workbook.save(file_path)

        print("Transaction added successfully!")
    except FileNotFoundError:
        print("Error: The file was not found. Check the file path.")
    except Exception as e:
        print(f"An error occured: {e}")



# File path to my spreadsheet
file_path = "C:/Users/akbba/OneDrive/Desktop/Finance Tracker.xlsx"

date = datetime.now().strftime("%d-%m-%Y")
description = "V-Bucks"
category = "Game"
amount = 10
transaction_type = "Expense"
recurring = "No"

append_transaction(file_path, date, description, category, amount, transaction_type, recurring)